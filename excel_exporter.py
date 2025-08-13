import logging
import os
import traceback
from collections import defaultdict
from datetime import date, datetime
from typing import Dict, List, Optional, DefaultDict, Any, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import (
    Alignment, Font, PatternFill, Border, Side, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import MergedCell

from data_processing import AdminStats, ServerStats, merge_duplicate_admins, fill_missing_roles
from utils import clean_sheet_name, normalize_admin_string

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
ALT_ROW_FILL = PatternFill(start_color="E6F0FF", end_color="E6F0FF", fill_type="solid")
MODERATOR_FILL = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
HIGHLIGHT_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
ADMIN_ONLY_FILL = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")
BORDER = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)


def clean_server_name(srv_name: str) -> str:
    prefix = "🤔┇ahelp-"
    if srv_name.startswith(prefix):
        srv_name = srv_name[len(prefix):]
    srv_name = srv_name.strip("_")
    srv_name = srv_name.replace("_", "-")
    return srv_name


def setup_worksheet_styles(workbook: Workbook) -> None:
    header_style = NamedStyle(name="header_style")
    header_style.font = HEADER_FONT
    header_style.fill = HEADER_FILL
    header_style.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_style.border = BORDER

    cell_style = NamedStyle(name="cell_style")
    cell_style.alignment = Alignment(horizontal="center", vertical="center")
    cell_style.border = BORDER

    alt_row_style = NamedStyle(name="alt_row_style")
    alt_row_style.fill = ALT_ROW_FILL
    alt_row_style.alignment = Alignment(horizontal="center", vertical="center")
    alt_row_style.border = BORDER

    admin_only_style = NamedStyle(name="admin_only_style")
    admin_only_style.fill = ADMIN_ONLY_FILL
    admin_only_style.alignment = Alignment(horizontal="center", vertical="center")
    admin_only_style.border = BORDER

    for style in [header_style, cell_style, alt_row_style, admin_only_style]:
        if style.name not in workbook.named_styles:
            workbook.add_named_style(style)


def adjust_column_widths(worksheet: Worksheet) -> None:
    column_widths = {
        'A': 30, 'B': 25, 'C': 15, 'D': 15, 'E': 15,
        'F': 18, 'G': 18, 'H': 18,
    }
    for i in range(9, 35):
        col_letter = get_column_letter(i)
        column_widths[col_letter] = 15
    for col_letter, width in column_widths.items():
        try:
            worksheet.column_dimensions[col_letter].width = width
        except Exception:
            pass


def apply_worksheet_formatting(
        worksheet: Worksheet, has_header: bool = True, highlight_moderators: bool = False,
        highlight_column: Optional[int] = None, admin_only_columns: Optional[List[int]] = None
) -> None:
    if has_header:
        for cell in worksheet[1]:
            if not isinstance(cell, MergedCell):
                cell.style = "header_style"

    start_row = 2 if has_header else 1
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=start_row), start=start_row):
        style_name = "alt_row_style" if row_idx % 2 == 0 else "cell_style"
        for cell in row:
            if isinstance(cell, MergedCell): continue
            cell.style = style_name
            if highlight_column and cell.column == highlight_column:
                cell.fill = HIGHLIGHT_FILL
            if admin_only_columns and cell.column in admin_only_columns:
                cell.fill = ADMIN_ONLY_FILL

        if highlight_moderators and len(row) > 1 and not isinstance(row[1], MergedCell):
            role_cell = row[1]
            role_value = str(role_cell.value).lower() if role_cell.value else ""
            if role_value and ("модератор" in role_value or "гейм-мастер" in role_value):
                for cell in row:
                    if not isinstance(cell, MergedCell):
                        cell.fill = MODERATOR_FILL


def add_metadata_to_worksheet(
        worksheet: Worksheet, title: str, description: str = "",
        data_range: str = "", server_name: str = ""
) -> None:
    worksheet.insert_rows(1, 4)
    worksheet.cell(1, 1, title)
    worksheet.cell(1, 1).font = Font(bold=True, size=16)
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    if server_name:
        worksheet.cell(2, 1, f"Сервер: {server_name}")
        worksheet.cell(2, 1).font = Font(bold=True)
    if description:
        worksheet.cell(3, 1, description)
    if data_range:
        worksheet.cell(3, 3, f"Диапазон дат: {data_range}")
    worksheet.cell(4, 1, f"Сгенерировано: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    worksheet.row_dimensions[5].height = 10


def write_df_to_excel_enhanced(
        df: pd.DataFrame, excel_filename: str, sheet_name: str, highlight_moderators: bool = False,
        sort_column: Optional[str] = None, ascending: bool = False,
        metadata: Optional[Dict[str, str]] = None, admin_only_columns: Optional[List[str]] = None,
) -> None:
    try:
        sheet_name = clean_sheet_name(sheet_name)
        if sort_column and sort_column in df.columns:
            df = df.sort_values(by=sort_column, ascending=ascending)

        if not os.path.exists(excel_filename):
            wb = Workbook()
            if "Sheet" in wb.sheetnames: del wb["Sheet"]
            setup_worksheet_styles(wb)
        else:
            wb = load_workbook(excel_filename)

        if sheet_name in wb.sheetnames: del wb[sheet_name]
        ws = wb.create_sheet(title=sheet_name)
        start_row = 1

        if metadata:
            add_metadata_to_worksheet(
                ws, title=metadata.get('title', sheet_name), description=metadata.get('description', ''),
                data_range=metadata.get('data_range', ''), server_name=metadata.get('server_name', '')
            )
            start_row = 6

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        highlight_column_idx = None
        if sort_column:
            try:
                highlight_column_idx = df.columns.get_loc(sort_column) + 1
            except KeyError:
                pass

        admin_only_column_indices = []
        if admin_only_columns:
            for col_name in admin_only_columns:
                try:
                    admin_only_column_indices.append(df.columns.get_loc(col_name) + 1)
                except KeyError:
                    pass

        apply_worksheet_formatting(
            ws, has_header=True, highlight_moderators=highlight_moderators,
            highlight_column=highlight_column_idx, admin_only_columns=admin_only_column_indices
        )

        if not df.empty:
            header_row = start_row
            last_row = start_row + len(df)
            last_col = len(df.columns)
            filter_range = f"A{header_row}:{get_column_letter(last_col)}{last_row}"
            try:
                ws.auto_filter.ref = filter_range
            except Exception as e:
                logging.warning(f"Не удалось установить автофильтр для диапазона {filter_range}: {e}")

        adjust_column_widths(ws)

        try:
            ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
        except Exception as e:
            logging.warning(f"Не удалось закрепить области: {e}")

        wb.save(excel_filename)
    except Exception as e:
        logging.error(f"Ошибка в write_df_to_excel_enhanced для листа '{sheet_name}': {e}")
        logging.error(traceback.format_exc())
        raise


def create_daily_ahelps_dataframe(daily_ahelps: DefaultDict[date, Dict[str, int]]) -> pd.DataFrame:
    if not daily_ahelps: return pd.DataFrame()
    all_dates = sorted(daily_ahelps.keys())
    normalized_daily_ahelps: DefaultDict[date, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for dt, admin_counts in daily_ahelps.items():
        for admin, count in admin_counts.items():
            normalized_daily_ahelps[dt][normalize_admin_string(admin)] += count
    all_admins = sorted({admin for daily_data in normalized_daily_ahelps.values() for admin in daily_data.keys()})
    data = []
    for admin in all_admins:
        total_ahelps = 0
        admin_row = [admin]
        for day in all_dates:
            count = normalized_daily_ahelps[day].get(admin, 0)
            total_ahelps += count
            admin_row.append(count)
        admin_row.append(total_ahelps)
        data.append(admin_row)
    df = pd.DataFrame(data, columns=["Administrator"] + [day.strftime("%Y-%m-%d") for day in all_dates] + ["Total"])
    return df.sort_values(by="Total", ascending=False)


def aggregate_daily_ahelps(servers_stats: Dict[str, ServerStats]) -> DefaultDict[date, Dict[str, int]]:
    global_daily_ahelps: DefaultDict[date, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for server_stats in servers_stats.values():
        for day, daily_data in server_stats["daily_ahelps"].items():
            for admin, count in daily_data.items():
                global_daily_ahelps[day][admin] += count
    return global_daily_ahelps


def create_hourly_ahelps_dataframe(hourly_ahelps: Dict[date, Dict[Any, Dict[str, int]]]) -> pd.DataFrame:
    rows = []
    for d, hours_data in hourly_ahelps.items():
        for h, vals in hours_data.items():
            try:
                hour = int(h)
            except (ValueError, TypeError):
                hour = 0
            total = vals.get("total", 0)
            processed = vals.get("processed", 0)
            percentage = round(processed / total * 100, 1) if total > 0 else 0
            rows.append([d.strftime("%Y-%m-%d"), hour, total, processed, percentage])
    if not rows:
        return pd.DataFrame(columns=["Date", "Hour", "Total Ahelps", "Processed", "Processed %"])
    df = pd.DataFrame(rows, columns=["Date", "Hour", "Total Ahelps", "Processed", "Processed %"])
    return df.sort_values(by=["Date", "Hour"])


def aggregate_hourly_ahelps(servers_stats: Dict[str, ServerStats]) -> Dict[date, Dict[int, Dict[str, int]]]:
    global_hourly_ahelps = defaultdict(lambda: defaultdict(lambda: {"total": 0, "processed": 0}))
    for server_stats in servers_stats.values():
        for d, hours_data in server_stats["hourly_ahelps"].items():
            for h, vals in hours_data.items():
                hour = int(h)
                global_hourly_ahelps[d][hour]["total"] += vals.get("total", 0)
                global_hourly_ahelps[d][hour]["processed"] += vals.get("processed", 0)
    return {d: dict(hours) for d, hours in global_hourly_ahelps.items()}


def create_summary_dataframe(
        global_admin_stats: Dict[str, AdminStats], global_chat_count: int, servers_stats: Dict[str, ServerStats]
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    server_data = []
    for server_name, stats in servers_stats.items():
        admin_count = len(stats["admin_stats"])
        ahelp_count = sum(a["ahelps"] for a in stats["admin_stats"].values())
        admin_only_ahelp_count = sum(a["admin_only_ahelps"] for a in stats["admin_stats"].values())
        mention_count = sum(a["mentions"] for a in stats["admin_stats"].values())
        admin_only_mention_count = sum(a["admin_only_mentions"] for a in stats["admin_stats"].values())
        moderator_count = sum(1 for a in stats["admin_stats"].values() if
                              "модератор" in a["role"].lower() or "гейм-мастер" in a["role"].lower())
        server_data.append([
            clean_server_name(server_name), stats["chat_count"], ahelp_count,
            admin_only_ahelp_count, admin_count, moderator_count, mention_count, admin_only_mention_count
        ])
    server_df = pd.DataFrame(server_data,
                             columns=["Server", "Chats", "Ahelps", "Admin Only Ahelps", "Admins", "Moderators",
                                      "Mentions", "Admin Only Mentions"])
    role_counts = defaultdict(lambda: defaultdict(int))
    for stats in global_admin_stats.values():
        role = normalize_admin_string(stats["role"])
        role_counts[role]["count"] += 1
        role_counts[role]["ahelps"] += stats["ahelps"]
        role_counts[role]["admin_only_ahelps"] += stats["admin_only_ahelps"]
    role_data = [[role, counts["count"], counts["ahelps"], counts["admin_only_ahelps"]] for role, counts in
                 role_counts.items()]
    role_df = pd.DataFrame(role_data, columns=["Role", "Admin Count", "Total Ahelps", "Admin Only Ahelps"])
    return server_df, role_df


def create_global_admins_dataframe(
        merged_global_stats: Dict[str, AdminStats], servers_stats: Dict[str, ServerStats]
) -> pd.DataFrame:
    admin_server_ahelps = defaultdict(dict)
    admin_server_admin_only_ahelps = defaultdict(dict)
    server_names = sorted(servers_stats.keys())
    cleaned_server_names = [clean_server_name(srv) for srv in server_names]

    for server_name, sstats in servers_stats.items():
        merged_server_stats = merge_duplicate_admins(sstats["admin_stats"])
        for admin, stats in merged_server_stats.items():
            normalized_admin = normalize_admin_string(admin)
            admin_server_ahelps[normalized_admin][clean_server_name(server_name)] = stats.get('ahelps', 0)
            admin_server_admin_only_ahelps[normalized_admin][clean_server_name(server_name)] = stats.get(
                'admin_only_ahelps', 0)

    global_data = []
    for admin, stats in merged_global_stats.items():
        row = [
            admin, stats['role'], stats.get('ahelps', 0), stats.get('mentions', 0),
            stats.get('sessions', 0), stats.get('admin_only_ahelps', 0),
            stats.get('admin_only_mentions', 0), stats.get('admin_only_sessions', 0)
        ]
        for srv in cleaned_server_names:
            row.append(admin_server_ahelps[admin].get(srv, 0))
        for srv in cleaned_server_names:
            row.append(admin_server_admin_only_ahelps[admin].get(srv, 0))
        global_data.append(row)

    global_columns = [
                         "Administrator", "Role", "Total Ahelps", "Mentions", "Sessions",
                         "Admin Only Ahelps", "Admin Only Mentions", "Admin Only Sessions"
                     ] + [f"{srv} Ahelps" for srv in cleaned_server_names] + [f"{srv} Admin Only" for srv in
                                                                              cleaned_server_names]
    return pd.DataFrame(global_data, columns=global_columns)


def save_all_data_to_excel(
        global_admin_stats: Dict[str, AdminStats], global_chat_count: int,
        servers_stats: Dict[str, ServerStats], df_global: pd.DataFrame
) -> None:
    try:
        excel_filename = os.getenv("EXCEL_FILENAME", "ahelp_stats.xlsx")
        if os.path.exists(excel_filename):
            try:
                os.remove(excel_filename)
            except OSError as e:
                logging.error(
                    f"Не удалось удалить существующий файл Excel '{excel_filename}': {e}. Возможно, он открыт.")
                return

        all_dates = [day for sstats in servers_stats.values() for day in sstats["daily_ahelps"]]
        date_range = ""
        if all_dates:
            min_date = min(all_dates).strftime("%Y-%m-%d")
            max_date = max(all_dates).strftime("%Y-%m-%d")
            date_range = f"{min_date} до {max_date}"

        server_summary_df, role_summary_df = create_summary_dataframe(global_admin_stats, global_chat_count,
                                                                      servers_stats)
        write_df_to_excel_enhanced(server_summary_df, excel_filename, "Summary",
                                   metadata={'title': 'Сводка по активности серверов', 'data_range': date_range},
                                   sort_column="Ahelps", ascending=False, admin_only_columns=["Admin Only Ahelps"])
        write_df_to_excel_enhanced(role_summary_df, excel_filename, "Roles_Summary",
                                   metadata={'title': 'Сводка по ролям администраторов', 'data_range': date_range},
                                   sort_column="Total Ahelps", ascending=False,
                                   admin_only_columns=["Admin Only Ahelps"])

        cleaned_server_names = [clean_server_name(srv) for srv in sorted(servers_stats.keys())]
        admin_only_cols = ["Admin Only Ahelps", "Admin Only Mentions",
                           "Admin Only Sessions"] + [f"{srv} Admin Only" for srv in cleaned_server_names]

        write_df_to_excel_enhanced(
            df_global, excel_filename, "Admins_Global", highlight_moderators=True,
            metadata={'title': 'Глобальная статистика администраторов', 'data_range': date_range},
            sort_column="Total Ahelps", ascending=False, admin_only_columns=admin_only_cols
        )

        df_moderators = df_global[
            df_global['Role'].str.lower().str.contains("модератор|гейм-мастер", na=False, regex=True)].copy()
        write_df_to_excel_enhanced(
            df_moderators, excel_filename, "Moderators",
            metadata={'title': 'Статистика модераторов', 'data_range': date_range},
            sort_column="Total Ahelps", ascending=False, admin_only_columns=admin_only_cols
        )

        for server_name, stats in servers_stats.items():
            clean_name = clean_server_name(server_name)
            daily_df = create_daily_ahelps_dataframe(stats["daily_ahelps"])
            if not daily_df.empty:
                write_df_to_excel_enhanced(daily_df, excel_filename, f"{clean_name}_Daily",
                                           metadata={'title': f'Ежедневные Ahelps - {clean_name}',
                                                     'data_range': date_range, 'server_name': clean_name},
                                           sort_column="Total", ascending=False)
            hourly_df = create_hourly_ahelps_dataframe(stats["hourly_ahelps"])
            if not hourly_df.empty:
                write_df_to_excel_enhanced(hourly_df, excel_filename, f"{clean_name}_Hourly",
                                           metadata={'title': f'Ежечасные Ahelps - {clean_name}',
                                                     'data_range': date_range, 'server_name': clean_name})

        global_daily_df = create_daily_ahelps_dataframe(aggregate_daily_ahelps(servers_stats))
        if not global_daily_df.empty:
            write_df_to_excel_enhanced(global_daily_df, excel_filename, "Daily_Global",
                                       metadata={'title': 'Глобальные ежедневные Ahelps', 'data_range': date_range},
                                       sort_column="Total", ascending=False)

        global_hourly_df = create_hourly_ahelps_dataframe(aggregate_hourly_ahelps(servers_stats))
        if not global_hourly_df.empty:
            write_df_to_excel_enhanced(global_hourly_df, excel_filename, "Hourly_Global",
                                       metadata={'title': 'Глобальные ежечасные Ahelps', 'data_range': date_range})

        logging.info(f"Все данные успешно сохранены в {excel_filename}")
    except Exception as e:
        logging.error(f"Критическая ошибка при сохранении всех данных в Excel: {e}")
        logging.error(traceback.format_exc())
        raise